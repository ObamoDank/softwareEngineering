import sqlite3
import openpyxl
import re
from openpyxl.utils import get_column_letter as gcl
from openpyxl.utils import column_index_from_string as cifs
import os.path as os

# Function take a list of years which it uses as the range and a list to test against that range. Function counts the number of times each year within the range appears within the list of years 
def sortYears(yearRange, yearList):
    sortedYears = {}
    for i in yearRange:
        sortedYears[str(i)] = 0
        for j in yearList:
            if i == int(j):
                sortedYears[str(i)] += 1
    return sortedYears

def givePercentage(n, total):
    x = round(n / total, 3) * 100
    if x == 0:
        return 'None'
    elif 0 < x < 1:
        return '<1%'
    else:
        return str(int(x)) + '%'

    
def calculateVerified(customers):
    total = len(customers)
    t = 0
    for i in customers:
        if i[1] == True:
            t += 1
    result = givePercentage(t, total)
    return result


def calculateGroup(customers):
    total = len(customers)
    amazon = 0
    kindle = 0
    other = 0
    for i in customers:
        if i[0] == 'Amazon Customer':
            amazon += 1
        elif i[0] == 'Kindle Customer':
            kindle += 1
        else:
            other += 1
    aPer = givePercentage(amazon, total)
    kPer = givePercentage(kindle, total)
    oPer = givePercentage(other, total)
    result = {'Amazon Customer': aPer, 'Kindle Customer': kPer, 'Other Customer': oPer}
    return result
    
connection = sqlite3.connect("product.db")
cursor = connection.cursor()

if os.exists('excel_review.xlsx'):
    wb = load_workbook('excel_review.xlsx')
else:
    wb = openpyxl.Workbook()
    
    
sheet = wb.active
sheet.title = "reviewsPerYear"

reviewsPerYear = []

# Store asin with title for later use
asin = {}

# Dictionary with brand -> asin -> year: reviews
brands = {}
years = []

reviews = cursor.execute("select i.asin, i.brand, i.title, r.dates from item i, review r where i.asin = r.asin").fetchall()

# Separate all of the data retrieved into structure as described above
for i in reviews:
#   Store asin with title for later use
    asin[i[0]] = i[2]
#   If brand doesn't exist, create new item in dictionary to store new brand
    if i[1] not in brands:
        brands[i[1]] = {}
#   Add each asin to corresponding brand as empty list to store a year for each entry
#   If item does not yet have a record of years, create space for record
    if i[0] not in brands[i[1]]:
        brands[i[1]][i[0]] = []
#   Search each date to extract year - add this year to the list
    yr = re.search(r"\d\d\d\d", i[3]).group()
    brands[i[1]][i[0]].append(yr)
#   Finally, add each unique year to a separate list to determine the time span from first review to last
    if yr not in years:
        years.append(yr)

# Sort years and convert from string to int
years.sort()
for i in range(len(years)):
    years[i] = int(years[i])

# We now have a a list of every year corresponding to each review stored for all items categorised by brands
# Next step is to determine the amount per year
for i in brands:
    for j in brands[i]:
        brands[i][j] = sortYears(years, brands[i][j])

# Start formatting cells in xlsx file
        
letter = 3
sheet['A1'] = 'brand'
sheet['B1'] = 'title'
for i in range(len(years)):
    sheet[gcl(letter) + '1'] = str(i + int(min(years)))
    letter += 1

# Insert all data into relevent cells
row = 2
for i in brands:
    for j in brands[i]:
        col = 3
        for k in brands[i][j]:
            sheet['A' + str(row)] = i
            sheet['B' + str(row)] = j
            sheet[gcl(col) + str(row)] = brands[i][j][k]
            col += 1
        row += 1
        
# Store Brands for the purpose of customaer analysis
cBrands = {}

customers = cursor.execute("select i.brand, r.name, r.verified from item i, review r where i.asin = r.asin").fetchall()
names = []
totalCustomers = len(customers)
totalAmazon = 0
totalKindle = 0
totalOther = 0

percentVerified = {}

# Sort all customers into their specific brands and store within dictionary
# Also counts total numbers of various customer groups
for i in customers:
    names.append(i[1])
    if i[0] not in cBrands:
        cBrands[i[0]] = []
    cBrands[i[0]].append([i[1], i[2]])
    if i[1] == 'Amazon Customer':
        totalAmazon += 1
    elif i[1] == 'Kindle Customer':
        totalKindle += 1
    else:
        totalOther += 1

for i in cBrands:
    percentVerified[i] = calculateVerified(cBrands[i])
    cBrands[i] = calculateGroup(cBrands[i])
    
# Next step is to store this data within a .xlsx file

wb.create_sheet(title = 'customers')
sheet = wb['customers']

# Format cells and add headers

sheet['A1'] = 'Brand'
sheet['B1'] = 'Verified Customers'
sheet['C1'] = 'Amazon Customers'
sheet['D1'] = 'Kindle Customers'
sheet['E1'] = 'Other Customers'

# Add data to cells. No need for intricate loops here as all data is known to developer before programming

row = 2
for i in cBrands:
    sheet['A' + str(row)] = i
    sheet['B' + str(row)] = percentVerified[i]
    sheet['C' + str(row)] = cBrands[i]['Amazon Customer']
    sheet['D' + str(row)] = cBrands[i]['Kindle Customer']
    sheet['E' + str(row)] = cBrands[i]['Other Customer']
    row += 1
    
wb.save('excel_review.xlsx')